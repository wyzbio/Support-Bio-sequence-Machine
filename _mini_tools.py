# -*- encoding: utf-8 -*-
# @Time     :   2023/03/11 15:23:57
# @Author   :   Yizheng Wang
# @E-mail   :   wyz020@126.com
# @Function :   保存一下常用的简单的IO代码片段，以供调用

import openpyxl
import time
import numpy as np


def save_list_to_excel(file_path, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in data:
        sheet.append(row)
    sheet.insert_rows(1)
    wb.save(file_path)


def save_list_to_txt(file_path, my_list):
    with open(file_path, "w") as f:
        for row in my_list:
            for col in row:
                f.write(str(col) + " ")
            f.write("\n")

# def save_list_to_txt(file_path, my_list):
#     with open(file_path, "w") as f:
#         if my_list and isinstance(my_list[0], list):  # 如果列表的第一个元素还是列表，我们认为它是二维的
#             for row in my_list:
#                 for col in row:
#                     f.write(str(col) + " ")
#                 f.write("\n")
#         else:  # 否则我们认为它是一维的
#             for elem in my_list:
#                 f.write(str(elem) + "\n")


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(list(row))
    return data


def read_fasta(file_path):
    with open(file_path, 'r') as f:
        lines = f.readlines()
        even_lines = [line.strip() for line in lines[1::2]]
        return even_lines


def read_txt(file_path):
    with open(file_path, 'r') as f:
        lines = f.readlines()
        matrix = [[float(num) for num in line.split()] for line in lines]
    return matrix


def read_row_txt(file_path):
    with open(file_path, 'r') as file:
        content_list = file.readlines()
    return content_list



def convert_npy_to_txt(file_path, output_file_path=None):
    # 从npy文件中读取数组
    my_array = np.load(file_path)
    # 将数组转换为列表
    my_list = my_array.tolist()
    # 如果有指定输出文件路径，则将列表保存到指定的文件中
    if output_file_path is not None:
        save_list_to_txt(output_file_path, my_list)
    else:
        # 否则，使用与输入文件相同的路径和文件名，但是将扩展名改为.txt
        txt_file_path = file_path.replace('.npy', '.txt')
        save_list_to_txt(txt_file_path, my_list)


def timer(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        print(f"函数{func.__name__}()的运行时间为：{end_time - start_time:.4f} 秒")
        return result
    return wrapper


class DataProcessor:

    def __init__(self):
        self.data = []

    def append_data(self, new_data):
        self.data.append(new_data)

    def average(self):
        if not self.data:
            return None
        return sum(self.data) / len(self.data)

    def maximum(self):
        if not self.data:
            return None
        return max(self.data)


if __name__ == "__main__":

    pass